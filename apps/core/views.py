"""
Core Views
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from django.contrib.auth import get_user_model
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from .models import OperationLog, SystemConfig, WeChatPayConfig, Permission, RolePermission
from .serializers import (
    UserSerializer, UserCreateSerializer, OperationLogSerializer, SystemConfigSerializer,
    WeChatPayConfigSerializer, WeChatPayConfigCreateSerializer,
    PermissionSerializer, RolePermissionSerializer, RolePermissionCreateSerializer
)
from .permissions import IsAdminUser

User = get_user_model()


def get_common_context():
    """获取所有页面共用的上下文数据"""
    from apps.maintenance.models import MaintenanceRequest
    return {
        'pending_maintenance_count': MaintenanceRequest.objects.filter(status='pending').count(),
        'unread_notifications': 3,  # 暂时固定，后续可从Notification表获取
    }


def log_operation(request, action, module, description):
    """
    记录操作日志

    Args:
        request: HttpRequest对象
        action: 操作类型（创建、更新、删除、登录、登出等）
        module: 模块名称（用户管理、社区管理、房产管理等）
        description: 操作描述
    """
    try:
        # 获取客户端IP地址
        ip_address = request.META.get('REMOTE_ADDR', '0.0.0.0')

        # 获取用户代理
        user_agent = request.META.get('HTTP_USER_AGENT', '')[:500]  # 限制长度

        # 创建日志记录
        OperationLog.objects.create(
            operator=request.user if request.user.is_authenticated else None,
            action=action,
            module=module,
            description=description,
            ip_address=ip_address,
            user_agent=user_agent
        )
    except Exception as e:
        # 日志记录失败不应影响主业务流程
        print(f'记录操作日志失败: {str(e)}')


@csrf_exempt
def index(request):
    """首页视图"""
    html_content = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>物业管理系统</title>
    <style>
        body {
            font-family: 'Microsoft YaHei', Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            margin: 0;
            padding: 0;
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
        }
        .container {
            background: white;
            border-radius: 10px;
            padding: 40px;
            box-shadow: 0 10px 25px rgba(0,0,0,0.2);
            text-align: center;
            max-width: 600px;
        }
        h1 {
            color: #333;
            margin-bottom: 20px;
        }
        p {
            color: #666;
            line-height: 1.6;
            margin-bottom: 15px;
        }
        .api-link {
            display: inline-block;
            margin: 10px;
            padding: 12px 24px;
            background: #667eea;
            color: white;
            text-decoration: none;
            border-radius: 5px;
            transition: background 0.3s;
        }
        .api-link:hover {
            background: #5568d3;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>欢迎使用物业管理系统</h1>
        <p>这是一个基于Django和Django REST Framework的物业管理系统API</p>
        <p>系统提供以下功能:</p>
        <ul style="text-align: left; color: #666;">
            <li>用户认证与权限管理</li>
            <li>社区信息管理</li>
            <li>房产信息管理</li>
            <li>缴费管理</li>
            <li>维修管理</li>
            <li>微信集成</li>
        </ul>
        <div style="margin-top: 30px;">
            <a href="/swagger/" class="api-link">API文档 (Swagger)</a>
            <a href="/redoc/" class="api-link">API文档 (ReDoc)</a>
            <a href="/admin/" class="api-link">管理后台</a>
        </div>
    </div>
</body>
</html>"""
    return HttpResponse(html_content, content_type='text/html; charset=utf-8')


class UserViewSet(viewsets.ModelViewSet):
    """用户管理视图集"""
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAdminUser]
    filterset_fields = ['role', 'is_active']
    search_fields = ['username', 'phone', 'first_name', 'last_name']

    def get_serializer_class(self):
        if self.action == 'create':
            return UserCreateSerializer
        return UserSerializer

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def me(self, request):
        """获取当前用户信息"""
        serializer = self.get_serializer(request.user)
        return Response(serializer.data)

    @action(detail=False, methods=['put', 'patch'], permission_classes=[IsAuthenticated])
    def update_profile(self, request):
        """更新当前用户信息"""
        serializer = self.get_serializer(request.user, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


class OperationLogViewSet(viewsets.ReadOnlyModelViewSet):
    """操作日志视图集"""
    queryset = OperationLog.objects.all()
    serializer_class = OperationLogSerializer
    permission_classes = [IsAdminUser]
    filterset_fields = ['action', 'module', 'operator']
    search_fields = ['description']


class SystemConfigViewSet(viewsets.ModelViewSet):
    """系统配置视图集"""
    queryset = SystemConfig.objects.all()
    serializer_class = SystemConfigSerializer
    permission_classes = [IsAdminUser]
    search_fields = ['key', 'description']


class WeChatPayConfigViewSet(viewsets.ModelViewSet):
    """微信支付配置视图集"""
    queryset = WeChatPayConfig.objects.all()
    permission_classes = [IsAdminUser]
    filterset_fields = ['account_type', 'is_active', 'is_default']
    search_fields = ['name', 'remarks']

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return WeChatPayConfigCreateSerializer
        return WeChatPayConfigSerializer

    @action(detail=True, methods=['post'])
    def set_default(self, request, pk=None):
        """设为默认配置"""
        config = self.get_object()
        # 取消其他默认配置
        WeChatPayConfig.objects.filter(is_default=True).update(is_default=False)
        # 设置当前为默认
        config.is_default = True
        config.save()
        return Response({'message': '已设为默认配置'}, status=status.HTTP_200_OK)


class PermissionViewSet(viewsets.ModelViewSet):
    """权限视图集"""
    queryset = Permission.objects.all()
    serializer_class = PermissionSerializer
    permission_classes = [IsAdminUser]
    filterset_fields = ['module']
    search_fields = ['name', 'code', 'description']


class RolePermissionViewSet(viewsets.ModelViewSet):
    """角色权限视图集"""
    queryset = RolePermission.objects.select_related('permission').all()
    permission_classes = [IsAdminUser]
    filterset_fields = ['role', 'permission__module']

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return RolePermissionCreateSerializer
        return RolePermissionSerializer

    @action(detail=False, methods=['get'])
    def by_role(self, request):
        """按角色获取权限"""
        role = request.query_params.get('role')
        if not role:
            return Response({'error': '请提供角色参数'}, status=status.HTTP_400_BAD_REQUEST)

        permissions = self.queryset.filter(role=role)
        serializer = self.get_serializer(permissions, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def bulk_update(self, request):
        """批量更新角色权限"""
        permissions_data = request.data.get('permissions', [])
        role = request.data.get('role')

        if not role:
            return Response({'error': '请提供角色参数'}, status=status.HTTP_400_BAD_REQUEST)

        # 删除该角色的所有权限
        RolePermission.objects.filter(role=role).delete()

        # 创建新权限
        created_permissions = []
        for perm_data in permissions_data:
            perm_data['role'] = role
            serializer = RolePermissionCreateSerializer(data=perm_data)
            if serializer.is_valid():
                created_permissions.append(serializer.save())

        return Response({
            'message': f'已更新角色 {role} 的权限',
            'count': len(created_permissions)
        }, status=status.HTTP_200_OK)


# ============================================
# 管理后台视图
# ============================================
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.forms import AuthenticationForm
from django.contrib import messages


def login_view(request):
    """登录视图"""
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                auth_login(request, user)
                # 记录登录日志
                log_operation(request, '登录', '用户管理', f'用户 {username} 登录系统')
                messages.info(request, f"欢迎回来，{username}！")
                return redirect('dashboard')
            else:
                # 记录失败的登录尝试
                log_operation(request, '登录失败', '用户管理', f'用户 {username} 登录失败：用户名或密码错误')
                messages.error(request, "用户名或密码错误")
        else:
            messages.error(request, "请输入有效的用户名和密码")
    else:
        form = AuthenticationForm()

    return render(request, 'registration/login.html', {'form': form})


@login_required
def logout_view(request):
    """登出视图"""
    from django.contrib.auth import logout
    username = request.user.username
    logout(request)
    # 记录登出日志
    log_operation(request, '登出', '用户管理', f'用户 {username} 登出系统')
    messages.info(request, "您已成功登出")
    return redirect('login')


@login_required
def dashboard(request):
    """仪表盘 - 数据概览"""
    from django.utils import timezone
    from django.db.models import Sum, Count, Q, F
    from apps.community.models import Community, Building
    from apps.property.models import Property
    from apps.payment.models import PaymentBill, PaymentRecord
    from apps.maintenance.models import MaintenanceRequest

    # 当前时间
    now = timezone.now()
    current_month_str = now.strftime('%Y-%m')  # 格式如 "2026-01"

    # 基础统计数据
    total_communities = Community.objects.count()
    total_buildings = Building.objects.count()
    total_households = Property.objects.count()

    # 本月缴费统计
    from django.db.models.functions import TruncMonth
    monthly_records = PaymentRecord.objects.filter(
        payment_time__year=now.year,
        payment_time__month=now.month
    ).aggregate(total=Sum('amount'))['total'] or 0

    # 本月应缴金额
    from decimal import Decimal
    monthly_bills = PaymentBill.objects.filter(
        billing_period__startswith=current_month_str
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # 收缴率
    collection_rate = (monthly_records / monthly_bills * 100) if monthly_bills > 0 else 0

    # 待处理报事
    pending_requests = MaintenanceRequest.objects.filter(status='pending').count()
    urgent_requests = MaintenanceRequest.objects.filter(
        status='pending',
        priority='high'
    ).count()

    # 逾期统计
    overdue_bills = PaymentBill.objects.filter(
        due_date__lt=now.date(),
        status__in=['unpaid', 'partial']
    )
    overdue_amount = overdue_bills.aggregate(total=Sum(F('amount') - F('paid_amount')))['total'] or Decimal('0')
    overdue_households = overdue_bills.values('owner').distinct().count()

    # 各小区收缴率
    community_stats = []
    for community in Community.objects.all():
        community_bills = PaymentBill.objects.filter(
            community=community,
            billing_period__startswith=current_month_str
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        community_records = PaymentRecord.objects.filter(
            bill__community=community,
            payment_time__year=now.year,
            payment_time__month=now.month
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        rate = (community_records / community_bills * 100) if community_bills > 0 else 0

        community_stats.append({
            'name': community.name,
            'rate': round(rate, 1),
            'bills_count': PaymentBill.objects.filter(
                community=community,
                billing_period__startswith=current_month_str
            ).count()
        })

    # 最近报事
    recent_requests = MaintenanceRequest.objects.select_related(
        'property', 'property__community', 'property__building'
    ).order_by('-created_at')[:5]

    # 待派单数量（用于侧边栏）
    pending_maintenance_count = pending_requests

    context = {
        'page_title': '数据概览',
        # 统计卡片数据
        'total_communities': total_communities,
        'total_households': total_households,
        'monthly_revenue': monthly_records,
        'collection_rate': round(collection_rate, 1),
        'pending_requests': pending_requests,
        'urgent_requests': urgent_requests,
        'overdue_amount': overdue_amount,
        'overdue_households': overdue_households,
        # 小区收缴率
        'community_stats': community_stats,
        # 最近报事
        'recent_requests': recent_requests,
        # 侧边栏数据
        'pending_maintenance_count': pending_maintenance_count,
        'unread_notifications': 3,
    }
    return render(request, 'admin/dashboard_full.html', context)


@login_required
def dashboard_stats_api(request):
    """仪表盘统计数据API"""
    from django.utils import timezone
    from django.db.models import Sum, Count, Q, F
    from apps.community.models import Community, Building
    from apps.property.models import Property
    from apps.payment.models import PaymentBill, PaymentRecord
    from apps.maintenance.models import MaintenanceRequest

    # 当前时间
    now = timezone.now()
    current_month_str = now.strftime('%Y-%m')

    # 基础统计数据
    total_communities = Community.objects.count()
    total_buildings = Building.objects.count()
    total_households = Property.objects.count()

    # 本月缴费统计
    monthly_records = PaymentRecord.objects.filter(
        payment_time__year=now.year,
        payment_time__month=now.month
    ).aggregate(total=Sum('amount'))['total'] or 0

    # 本月应缴金额
    from decimal import Decimal
    monthly_bills = PaymentBill.objects.filter(
        billing_period__startswith=current_month_str
    ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

    # 收缴率
    collection_rate = (monthly_records / monthly_bills * 100) if monthly_bills > 0 else 0

    # 待处理报事
    pending_requests = MaintenanceRequest.objects.filter(status='pending').count()
    urgent_requests = MaintenanceRequest.objects.filter(
        status='pending',
        priority='high'
    ).count()

    # 逾期统计
    overdue_bills = PaymentBill.objects.filter(
        due_date__lt=now.date(),
        status__in=['unpaid', 'partial']
    )
    overdue_amount = overdue_bills.aggregate(total=Sum(F('amount') - F('paid_amount')))['total'] or Decimal('0')
    overdue_households = overdue_bills.values('property__ownerproperty__owner').distinct().count()

    # 各小区收缴率
    community_stats = []
    for community in Community.objects.all():
        community_bills = PaymentBill.objects.filter(
            community=community,
            billing_period__startswith=current_month_str
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        community_records = PaymentRecord.objects.filter(
            bill__community=community,
            payment_time__year=now.year,
            payment_time__month=now.month
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        rate = (community_records / community_bills * 100) if community_bills > 0 else 0

        community_stats.append({
            'name': community.name,
            'rate': round(rate, 1),
            'bills_count': PaymentBill.objects.filter(
                community=community,
                billing_period__startswith=current_month_str
            ).count()
        })

    # 最近报事
    recent_requests = MaintenanceRequest.objects.select_related(
        'property', 'property__community', 'property__building'
    ).order_by('-created_at')[:5]

    # 待派单数量（用于侧边栏）
    pending_maintenance_count = pending_requests

    # 报事统计（按类别）
    requests_by_category = []
    categories = {
        'electric': '电工',
        'plumbing': '水工',
        'civil': '土木',
        'elevator': '电梯',
        'cleaning': '清洁',
        'security': '安保',
        'other': '其他'
    }

    for category_key, category_name in categories.items():
        count = MaintenanceRequest.objects.filter(
            category=category_key,
            created_at__month=now.month,
            created_at__year=now.year
        ).count()

        requests_by_category.append({
            'category': category_name,
            'count': count,
            'key': category_key
        })

    return JsonResponse({
        'success': True,
        'data': {
            # 统计卡片数据
            'total_communities': total_communities,
            'total_buildings': total_buildings,
            'total_households': total_households,
            'monthly_revenue': str(monthly_records),
            'collection_rate': round(collection_rate, 1),
            'pending_requests': pending_requests,
            'urgent_requests': urgent_requests,
            'overdue_amount': str(overdue_amount),
            'overdue_households': overdue_households,
            # 小区收缴率
            'community_stats': community_stats,
            # 最近报事
            'recent_requests': [
                {
                    'id': str(req.id),
                    'request_number': req.request_number,
                    'category': req.get_category_display(),
                    'priority': req.get_priority_display(),
                    'status': req.get_status_display(),
                    'created_at': req.created_at.strftime('%Y-%m-%d %H:%M'),
                    'property_address': req.property.full_address if req.property else '',
                    'reporter': req.reporter
                }
                for req in recent_requests
            ],
            # 报事类别统计
            'requests_by_category': requests_by_category,
            # 侧边栏数据
            'pending_maintenance_count': pending_maintenance_count,
        }
    })


@login_required
def community_list(request):
    """小区管理（带搜索和筛选）"""
    from apps.community.models import Community, Building
    from django.db.models import Q

    # 获取搜索和筛选参数
    search_query = request.GET.get('search', '').strip()

    # 小区查询集
    communities_queryset = Community.objects.all()

    # 小区搜索逻辑：搜索小区名称、地址、开发商、物业公司
    if search_query:
        communities_queryset = communities_queryset.filter(
            Q(name__icontains=search_query) |
            Q(address__icontains=search_query) |
            Q(developer__icontains=search_query) |
            Q(property_company__icontains=search_query)
        )

    communities = communities_queryset.order_by('created_at')

    # 楼栋查询集
    buildings_queryset = Building.objects.select_related('community').all()

    # 楼栋搜索逻辑：搜索楼栋号、小区名称
    if search_query:
        buildings_queryset = buildings_queryset.filter(
            Q(name__icontains=search_query) |
            Q(community__name__icontains=search_query)
        )

    buildings = buildings_queryset.order_by('community__name', 'name')

    context = {
        'communities': communities,
        'buildings': buildings,
        'search_query': search_query,
    }
    context.update(get_common_context())
    return render(request, 'admin/community.html', context)


@login_required
def property_list(request):
    """房产管理（带分页和搜索）"""
    from apps.property.models import Property, Owner, Tenant
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.db.models import Q, Prefetch
    from apps.property.models import OwnerProperty
    from apps.community.models import Community

    # 获取分页参数
    page = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 20)

    # 验证page_size
    allowed_page_sizes = [10, 20, 50, 100]
    try:
        page_size = int(page_size)
        if page_size not in allowed_page_sizes:
            page_size = 20
    except (ValueError, TypeError):
        page_size = 20

    # 获取搜索和筛选参数
    search_query = request.GET.get('search', '').strip()
    filter_community = request.GET.get('community', '')
    filter_property_type = request.GET.get('property_type', '')
    filter_status = request.GET.get('status', '')

    # 房产查询集（预加载业主信息）
    properties_queryset = Property.objects.select_related(
        'community', 'building'
    ).prefetch_related(
        Prefetch(
            'owners',
            queryset=OwnerProperty.objects.select_related('owner').order_by('-is_primary', '-created_at')
        )
    )

    # 搜索逻辑：搜索房号、业主姓名、联系电话
    if search_query:
        # 尝试解析搜索词是否为数字房号（如2704）
        # 房号格式：floor + room_number（如 27 + 04 = 2704）
        floor_room_match = None
        if search_query.isdigit() and len(search_query) >= 3:
            # 根据数字长度智能解析
            # 3位数(501): floor=第1位, room=后2位
            # 4位数(2704): floor=前2位, room=后2位
            # 5位数(10001): floor=前3位, room=后2位
            query_len = len(search_query)
            room_part_len = 2  # 房间号通常是2位数

            if query_len == 3:
                # 501 → floor=5, room="01"
                floor_num = int(search_query[0])
                room_suffix = search_query[1:]
                floor_room_match = (floor_num, room_suffix)
            elif query_len == 4:
                # 2704 → floor=27, room="04"
                floor_num = int(search_query[:2])
                room_suffix = search_query[2:]
                floor_room_match = (floor_num, room_suffix)
            elif query_len == 5:
                # 10001 → floor=100, room="01"
                floor_num = int(search_query[:3])
                room_suffix = search_query[3:]
                floor_room_match = (floor_num, room_suffix)
            # 更多位数的房号可以继续扩展

        # 构建查询条件
        queries = Q()

        # 1. 搜索楼栋名称（如"4号楼"）
        queries |= Q(building__name__icontains=search_query)

        # 2. 搜索纯房间号（如"04"）
        queries |= Q(room_number__icontains=search_query)

        # 3. 如果匹配到楼层+房间号模式（如2704 → floor=27, room_number__startswith="04"）
        if floor_room_match:
            floor_num, room_suffix = floor_room_match
            queries |= Q(floor=floor_num, room_number__startswith=room_suffix)

        # 应用房号查询
        properties_queryset = properties_queryset.filter(queries)

        # 4. 同时搜索业主姓名和电话
        owner_ids = Owner.objects.filter(
            Q(name__icontains=search_query) |
            Q(phone__icontains=search_query)
        ).values_list('id', flat=True)

        # 获取这些业主关联的房产ID
        if owner_ids:
            property_ids_from_owners = OwnerProperty.objects.filter(
                owner_id__in=owner_ids
            ).values_list('property_id', flat=True)

            # 使用OR连接查询：房号匹配 OR 业主匹配
            properties_queryset = properties_queryset | Property.objects.filter(
                id__in=property_ids_from_owners
            ).select_related(
                'community', 'building'
            ).prefetch_related(
                Prefetch(
                    'owners',
                    queryset=OwnerProperty.objects.select_related('owner').order_by('-is_primary', '-created_at')
                )
            )

    # 应用筛选条件
    if filter_community:
        properties_queryset = properties_queryset.filter(community_id=filter_community)
    if filter_property_type:
        properties_queryset = properties_queryset.filter(property_type=filter_property_type)
    if filter_status:
        properties_queryset = properties_queryset.filter(status=filter_status)

    properties_queryset = properties_queryset.order_by('building', 'floor', 'room_number').distinct()
    property_paginator = Paginator(properties_queryset, page_size)

    try:
        properties_page = property_paginator.page(page)
    except PageNotAnInteger:
        properties_page = property_paginator.page(1)
    except EmptyPage:
        properties_page = property_paginator.page(property_paginator.num_pages)

    # 业主分页（预加载房产信息）- 添加搜索和筛选
    owners_queryset = Owner.objects.prefetch_related(
        'owners__property__building'
    ).order_by('-created_at')

    # 获取业主筛选参数
    owner_filter_community = request.GET.get('owner_community', '')
    owner_filter_verified = request.GET.get('owner_verified', '')

    # 业主搜索逻辑：搜索姓名、手机号、房产
    if search_query:
        # 搜索业主姓名和手机号
        owners_queryset = owners_queryset.filter(
            Q(name__icontains=search_query) |
            Q(phone__icontains=search_query)
        )

        # 搜索房产（楼栋、房号）
        # 先获取匹配的房产ID
        from django.db.models import Q
        property_queries = Q(building__name__icontains=search_query) | Q(room_number__icontains=search_query)

        # 尝试解析数字房号
        if search_query.isdigit() and len(search_query) >= 3:
            query_len = len(search_query)
            if query_len == 3:
                floor_num = int(search_query[0])
                room_suffix = search_query[1:]
                property_queries |= Q(floor=floor_num, room_number__startswith=room_suffix)
            elif query_len == 4:
                floor_num = int(search_query[:2])
                room_suffix = search_query[2:]
                property_queries |= Q(floor=floor_num, room_number__startswith=room_suffix)
            elif query_len == 5:
                floor_num = int(search_query[:3])
                room_suffix = search_query[3:]
                property_queries |= Q(floor=floor_num, room_number__startswith=room_suffix)

        # 获取匹配的房产ID
        matching_property_ids = Property.objects.filter(property_queries).values_list('id', flat=True)

        # 获取这些房产关联的业主ID
        if matching_property_ids:
            owner_ids_from_properties = OwnerProperty.objects.filter(
                property_id__in=matching_property_ids
            ).values_list('owner_id', flat=True)

            # 使用OR连接：姓名/电话匹配 OR 房产匹配
            owners_queryset = owners_queryset | Owner.objects.filter(
                id__in=owner_ids_from_properties
            ).prefetch_related(
                'owners__property__building'
            )

    # 应用业主筛选条件
    if owner_filter_community:
        # 筛选拥有指定小区房产的业主
        owner_ids_in_community = OwnerProperty.objects.filter(
            property__community_id=owner_filter_community
        ).values_list('owner_id', flat=True)
        owners_queryset = owners_queryset.filter(id__in=owner_ids_in_community)

    if owner_filter_verified:
        if owner_filter_verified == 'verified':
            owners_queryset = owners_queryset.filter(is_verified=True)
        elif owner_filter_verified == 'unverified':
            owners_queryset = owners_queryset.filter(is_verified=False)

    owners_queryset = owners_queryset.order_by('-created_at').distinct()
    owner_paginator = Paginator(owners_queryset, page_size)

    try:
        owners_page = owner_paginator.page(page)
    except PageNotAnInteger:
        owners_page = owner_paginator.page(1)
    except EmptyPage:
        owners_page = owner_paginator.page(owner_paginator.num_pages)

    # 获取租户筛选参数
    tenant_filter_community = request.GET.get('tenant_community', '')
    tenant_filter_status = request.GET.get('tenant_status', '')

    # 租户分页 - 添加搜索
    tenants_queryset = Tenant.objects.select_related('property__community', 'property__building').order_by('-created_at')

    # 租户小区筛选逻辑
    if tenant_filter_community:
        tenants_queryset = tenants_queryset.filter(property__community_id=tenant_filter_community)

    # 租户状态筛选逻辑
    if tenant_filter_status:
        from django.utils import timezone
        today = timezone.now().date()
        if tenant_filter_status == 'active':
            # 租赁中：lease_end >= 今天
            tenants_queryset = tenants_queryset.filter(lease_end__gte=today)
        elif tenant_filter_status == 'expired':
            # 已到期：lease_end < 今天
            tenants_queryset = tenants_queryset.filter(lease_end__lt=today)

    # 租户搜索逻辑：搜索姓名、手机号、房产
    if search_query:
        # 搜索租户姓名和手机号
        tenants_queryset = tenants_queryset.filter(
            Q(name__icontains=search_query) |
            Q(phone__icontains=search_query)
        )

        # 搜索房产（楼栋、房号）
        property_queries = Q(property__building__name__icontains=search_query) | Q(property__room_number__icontains=search_query)

        # 尝试解析数字房号
        if search_query.isdigit() and len(search_query) >= 3:
            query_len = len(search_query)
            if query_len == 3:
                floor_num = int(search_query[0])
                room_suffix = search_query[1:]
                property_queries |= Q(property__floor=floor_num, property__room_number__startswith=room_suffix)
            elif query_len == 4:
                floor_num = int(search_query[:2])
                room_suffix = search_query[2:]
                property_queries |= Q(property__floor=floor_num, property__room_number__startswith=room_suffix)
            elif query_len == 5:
                floor_num = int(search_query[:3])
                room_suffix = search_query[3:]
                property_queries |= Q(property__floor=floor_num, property__room_number__startswith=room_suffix)

        # 使用OR连接：姓名/电话匹配 OR 房产匹配
        # 注意：需要再次应用小区和状态筛选
        base_queryset = tenants_queryset
        tenants_queryset = tenants_queryset | Tenant.objects.select_related('property__community', 'property__building').filter(property_queries)

        # 如果有小区筛选，需要对合并后的查询集再次应用
        if tenant_filter_community:
            tenants_queryset = tenants_queryset.filter(property__community_id=tenant_filter_community)
        # 如果有状态筛选，需要对合并后的查询集再次应用
        if tenant_filter_status:
            from django.utils import timezone
            today = timezone.now().date()
            if tenant_filter_status == 'active':
                tenants_queryset = tenants_queryset.filter(lease_end__gte=today)
            elif tenant_filter_status == 'expired':
                tenants_queryset = tenants_queryset.filter(lease_end__lt=today)

    tenants_queryset = tenants_queryset.order_by('-created_at').distinct()
    tenant_paginator = Paginator(tenants_queryset, page_size)

    try:
        tenants_page = tenant_paginator.page(page)
    except PageNotAnInteger:
        tenants_page = tenant_paginator.page(1)
    except EmptyPage:
        tenants_page = tenant_paginator.page(tenant_paginator.num_pages)

    # 获取所有小区用于筛选下拉框
    communities = Community.objects.all().order_by('name')

    context = {
        'properties': properties_page,
        'owners': owners_page,
        'tenants': tenants_page,
        'page_size': page_size,
        'allowed_page_sizes': allowed_page_sizes,
        # 搜索和筛选参数
        'search_query': search_query,
        'communities': communities,
        'owner_filter_community': owner_filter_community,
        'owner_filter_verified': owner_filter_verified,
        'tenant_filter_community': tenant_filter_community,
        'tenant_filter_status': tenant_filter_status,
        # 额外的分页信息
        'property_total': property_paginator.count,
        'owner_total': owner_paginator.count,
        'tenant_total': tenant_paginator.count,
    }
    context.update(get_common_context())
    return render(request, 'admin/property.html', context)


@login_required
def payment_list(request):
    """缴费管理（带分页和搜索）"""
    from apps.payment.models import PaymentBill, FeeStandard, PaymentRecord
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    from django.db.models import Q
    from apps.community.models import Community

    # 获取分页参数
    page = request.GET.get('page', 1)
    page_size = request.GET.get('page_size', 10)

    # 验证page_size
    allowed_page_sizes = [10, 20, 50, 100]
    try:
        page_size = int(page_size)
        if page_size not in allowed_page_sizes:
            page_size = 10
    except (ValueError, TypeError):
        page_size = 10

    # 获取搜索和筛选参数
    search_query = request.GET.get('search', '').strip()

    # ========== 账单列表 ==========
    # 获取账单筛选参数
    bill_filter_community = request.GET.get('community', '')
    bill_filter_fee_type = request.GET.get('fee_type', '')
    bill_filter_status = request.GET.get('status', '')

    # 构建账单查询集
    bills_queryset = PaymentBill.objects.select_related(
        'property_unit', 'property_unit__community', 'property_unit__building', 'owner', 'community'
    )

    # 账单搜索逻辑 - 增强版：支持房号、业主、账单编号、应缴金额
    if search_query:
        # 1. 搜索账单编号
        queries = Q(bill_number__icontains=search_query)

        # 2. 搜索业主姓名
        queries |= Q(owner__name__icontains=search_query)

        # 3. 搜索应缴金额（尝试将搜索词解析为数字）
        try:
            amount = float(search_query)
            queries |= Q(amount=amount)
        except (ValueError, TypeError):
            pass

        # 4. 搜索房号（智能解析，参考房产管理的实现）
        # 尝试解析搜索词是否为数字房号（如502, 2704等）
        if search_query.isdigit() and len(search_query) >= 3:
            # 房号格式：floor + room_number（如 5 + 02 = 502）
            floor_room_match = None
            query_len = len(search_query)
            room_part_len = 2  # 房间号通常是2位数

            if query_len == 3:
                # 501 → floor=5, room="01"
                floor_num = int(search_query[0])
                room_suffix = search_query[1:]
                floor_room_match = (floor_num, room_suffix)
            elif query_len == 4:
                # 2704 → floor=27, room="04"
                floor_num = int(search_query[:2])
                room_suffix = search_query[2:]
                floor_room_match = (floor_num, room_suffix)
            elif query_len == 5:
                # 10001 → floor=100, room="01"
                floor_num = int(search_query[:3])
                room_suffix = search_query[3:]
                floor_room_match = (floor_num, room_suffix)

            # 如果匹配到楼层+房间号模式，添加到查询条件
            if floor_room_match:
                floor_num, room_suffix = floor_room_match
                queries |= Q(property_unit__floor=floor_num, property_unit__room_number__startswith=room_suffix)

        # 5. 搜索纯房间号（如"02"或"502"）
        queries |= Q(property_unit__room_number__icontains=search_query)

        # 6. 搜索楼栋名称（如"4号楼"）
        queries |= Q(property_unit__building__name__icontains=search_query)

        # 应用所有查询条件
        bills_queryset = bills_queryset.filter(queries)

    # 应用账单筛选条件
    if bill_filter_community:
        bills_queryset = bills_queryset.filter(community_id=bill_filter_community)
    if bill_filter_fee_type:
        bills_queryset = bills_queryset.filter(fee_type=bill_filter_fee_type)
    if bill_filter_status:
        bills_queryset = bills_queryset.filter(status=bill_filter_status)

    bills_queryset = bills_queryset.order_by('-created_at').distinct()
    bill_paginator = Paginator(bills_queryset, page_size)

    try:
        bills_page = bill_paginator.page(page)
    except PageNotAnInteger:
        bills_page = bill_paginator.page(1)
    except EmptyPage:
        bills_page = bill_paginator.page(bill_paginator.num_pages)

    # ========== 费用标准 ==========
    # 获取费用标准筛选参数
    standard_filter_community = request.GET.get('community', '')
    standard_filter_fee_type = request.GET.get('fee_type', '')

    # 构建费用标准查询集
    standards_queryset = FeeStandard.objects.select_related('community')

    # 应用费用标准筛选条件
    if standard_filter_community:
        standards_queryset = standards_queryset.filter(community_id=standard_filter_community)
    if standard_filter_fee_type:
        standards_queryset = standards_queryset.filter(fee_type=standard_filter_fee_type)

    standards_queryset = standards_queryset.order_by('community__name', 'fee_type')
    standard_paginator = Paginator(standards_queryset, page_size)

    try:
        standards_page = standard_paginator.page(page)
    except PageNotAnInteger:
        standards_page = standard_paginator.page(1)
    except EmptyPage:
        standards_page = standard_paginator.page(standard_paginator.num_pages)

    # ========== 缴费记录 ==========
    # 获取缴费记录筛选参数
    record_filter_community = request.GET.get('community', '')
    record_filter_payment_method = request.GET.get('payment_method', '')

    # 构建缴费记录查询集
    records_queryset = PaymentRecord.objects.select_related(
        'bill', 'bill__owner', 'bill__property_unit', 'bill__property_unit__community', 'bill__property_unit__building'
    )

    # 缴费记录搜索逻辑
    if search_query:
        records_queryset = records_queryset.filter(
            Q(transaction_id__icontains=search_query) |
            Q(bill__property_unit__room_number__icontains=search_query) |
            Q(bill__owner__name__icontains=search_query)
        )

    # 应用缴费记录筛选条件
    if record_filter_community:
        records_queryset = records_queryset.filter(bill__community_id=record_filter_community)
    if record_filter_payment_method:
        records_queryset = records_queryset.filter(payment_method=record_filter_payment_method)

    records_queryset = records_queryset.order_by('-payment_time')
    record_paginator = Paginator(records_queryset, page_size)

    try:
        records_page = record_paginator.page(page)
    except PageNotAnInteger:
        records_page = record_paginator.page(1)
    except EmptyPage:
        records_page = record_paginator.page(record_paginator.num_pages)

    # 获取所有小区用于筛选下拉框
    communities = Community.objects.all().order_by('name')

    context = {
        'bills': bills_page,
        'standards': standards_page,
        'records': records_page,
        'page_size': page_size,
        'allowed_page_sizes': allowed_page_sizes,
        # 搜索和筛选参数
        'search_query': search_query,
        'communities': communities,
        # 额外的分页信息
        'bill_total': bill_paginator.count,
        'standard_total': standard_paginator.count,
        'record_total': record_paginator.count,
    }
    context.update(get_common_context())
    return render(request, 'admin/payment.html', context)


@login_required
def maintenance_list(request):
    """报事管理"""
    from apps.maintenance.models import MaintenanceRequest
    from apps.community.models import Community
    from django.db.models import Count, Q
    from django.utils import timezone
    from django.core.paginator import Paginator

    # 获取所有小区（用于筛选下拉框）
    communities = Community.objects.all()

    # 获取筛选参数
    search_query = request.GET.get('search', '')
    community_id = request.GET.get('community', '')
    category = request.GET.get('category', '')
    priority = request.GET.get('priority', '')
    status = request.GET.get('status', '')

    # 基础查询
    requests = MaintenanceRequest.objects.select_related(
        'property', 'property__community', 'property__building', 'community'
    ).all()

    # 应用筛选
    if search_query:
        requests = requests.filter(
            Q(request_number__icontains=search_query) |
            Q(property__room_number__icontains=search_query) |
            Q(reporter__icontains=search_query)
        )

    if community_id:
        requests = requests.filter(community_id=community_id)

    if category:
        requests = requests.filter(category=category)

    if priority:
        requests = requests.filter(priority=priority)

    if status:
        requests = requests.filter(status=status)

    # 按状态分类统计（基于筛选后的结果）
    pending_requests = list(requests.filter(status='pending'))
    assigned_requests = list(requests.filter(status='assigned'))
    processing_requests = list(requests.filter(status='processing'))
    completed_requests = list(requests.filter(status='completed')[:10])  # 只显示最近10条

    # 统计数据（基于筛选后的结果）
    stats_by_status = {
        'pending': requests.filter(status='pending').count(),
        'assigned': requests.filter(status='assigned').count(),
        'processing': requests.filter(status='processing').count(),
        'completed': requests.filter(status='completed').count(),
    }

    # 按类别统计
    now = timezone.now()
    current_month = now.month
    current_year = now.year

    # 本月报事统计（按类别）
    from django.db.models.functions import TruncMonth
    monthly_requests = requests.filter(
        created_at__year=current_year,
        created_at__month=current_month
    )

    category_stats = {}
    total_monthly = monthly_requests.count()

    for category_choice in ['electric', 'plumbing', 'civil', 'elevator', 'cleaning', 'security', 'other']:
        count = monthly_requests.filter(category=category_choice).count()
        if count > 0:
            category_name = dict(MaintenanceRequest.CATEGORY_CHOICES).get(category_choice, category_choice)
            percentage = (count / total_monthly * 100) if total_monthly > 0 else 0
            category_stats[category_name] = {
                'count': count,
                'percentage': round(percentage, 1)
            }

    # 按优先级统计（基于筛选后的结果）
    priority_stats = {
        'high': requests.filter(priority='high').count(),
        'medium': requests.filter(priority='medium').count(),
        'low': requests.filter(priority='low').count(),
    }

    # 分页
    page_size = int(request.GET.get('page_size', 20))
    page_number = request.GET.get('page', 1)
    paginator = Paginator(requests, page_size)
    requests_page = paginator.get_page(page_number)

    context = {
        'requests': requests_page,
        'request_total': requests.count(),
        'page_size': page_size,
        # 小区列表
        'communities': communities,
        # 搜索和筛选参数
        'search_query': search_query,
        # 看板数据
        'pending_requests': pending_requests,
        'assigned_requests': assigned_requests,
        'processing_requests': processing_requests,
        'completed_requests': completed_requests,
        # 统计数据
        'stats_by_status': stats_by_status,
        'category_stats': category_stats,
        'priority_stats': priority_stats,
        'total_monthly': total_monthly,
    }
    context.update(get_common_context())
    return render(request, 'admin/maintenance.html', context)


@login_required
def user_list(request):
    """用户管理"""
    from django.db.models import Count, Q

    users = User.objects.all()

    # 统计数据
    total_users = users.count()
    admin_count = users.filter(role='admin').count()
    active_users_count = users.filter(is_active=True).count()
    owner_count = users.filter(role='owner').count()
    tenant_count = users.filter(role='tenant').count()
    staff_count = users.filter(role='staff').count()

    context = {
        'users': users,
        'total_users': total_users,
        'admin_count': admin_count,
        'active_users_count': active_users_count,
        'owner_count': owner_count,
        'tenant_count': tenant_count,
        'staff_count': staff_count,
    }
    context.update(get_common_context())
    return render(request, 'admin/users.html', context)


@login_required
def settings_list(request):
    """系统设置"""
    configs = SystemConfig.objects.all()

    context = {
        'configs': configs,
    }
    context.update(get_common_context())
    return render(request, 'admin/settings.html', context)


@login_required
def log_list(request):
    """操作日志"""
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

    # 获取筛选参数
    search_query = request.GET.get('search', '')
    module_filter = request.GET.get('module', '')
    action_filter = request.GET.get('action', '')
    date_filter = request.GET.get('date', '')

    # 获取所有日志
    logs_queryset = OperationLog.objects.select_related('operator').all()

    # 应用筛选条件
    if search_query:
        logs_queryset = logs_queryset.filter(
            Q(operator__username__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    if module_filter:
        logs_queryset = logs_queryset.filter(module=module_filter)

    if action_filter:
        logs_queryset = logs_queryset.filter(action=action_filter)

    if date_filter:
        logs_queryset = logs_queryset.filter(created_at__date=date_filter)

    # 分页
    paginator = Paginator(logs_queryset, 50)  # 每页50条
    page = request.GET.get('page', 1)

    try:
        logs = paginator.page(page)
    except PageNotAnInteger:
        logs = paginator.page(1)
    except EmptyPage:
        logs = paginator.page(paginator.num_pages)

    # 计算统计数据
    from django.utils import timezone
    from django.db.models import Count
    from datetime import timedelta

    today = timezone.now().date()
    week_ago = today - timedelta(days=7)

    today_count = OperationLog.objects.filter(created_at__date=today).count()
    week_count = OperationLog.objects.filter(created_at__date__gte=week_ago).count()
    active_users = OperationLog.objects.filter(
        created_at__date__gte=week_ago
    ).values('operator').distinct().count()

    context = {
        'logs': logs,
        'today_count': today_count,
        'week_count': week_count,
        'active_users': active_users,
        'module_filter': module_filter,
        'action_filter': action_filter,
        'date_filter': date_filter,
        'search_query': search_query,
    }
    context.update(get_common_context())
    return render(request, 'admin/logs.html', context)


@login_required
def export_logs(request):
    """导出操作日志为Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.http import HttpResponse
    from datetime import datetime

    # 记录导出操作
    log_operation(request, '导出', '系统管理', '导出操作日志')

    # 获取筛选参数
    search_query = request.GET.get('search', '')
    module_filter = request.GET.get('module', '')
    action_filter = request.GET.get('action', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')

    # 获取日志数据
    logs_queryset = OperationLog.objects.select_related('operator').all()

    # 应用筛选条件
    if search_query:
        logs_queryset = logs_queryset.filter(
            Q(operator__username__icontains=search_query) |
            Q(description__icontains=search_query)
        )

    if module_filter:
        logs_queryset = logs_queryset.filter(module=module_filter)

    if action_filter:
        logs_queryset = logs_queryset.filter(action=action_filter)

    if date_from:
        logs_queryset = logs_queryset.filter(created_at__date__gte=date_from)

    if date_to:
        logs_queryset = logs_queryset.filter(created_at__date__lte=date_to)

    logs = logs_queryset.order_by('-created_at')[:10000]  # 最多导出10000条

    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '操作日志'

    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 设置列标题
    headers = ['时间', '操作人', '角色', '操作类型', '模块', '操作描述', 'IP地址', '用户代理']
    ws.append(headers)

    # 设置列宽
    ws.column_dimensions['A'].width = 20  # 时间
    ws.column_dimensions['B'].width = 15  # 操作人
    ws.column_dimensions['C'].width = 12  # 角色
    ws.column_dimensions['D'].width = 12  # 操作类型
    ws.column_dimensions['E'].width = 15  # 模块
    ws.column_dimensions['F'].width = 40  # 操作描述
    ws.column_dimensions['G'].width = 15  # IP地址
    ws.column_dimensions['H'].width = 30  # 用户代理

    # 设置标题行样式
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 填充数据
    for row_num, log in enumerate(logs, 2):
        ws.cell(row=row_num, column=1, value=log.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row_num, column=2, value=log.operator.username if log.operator else '系统')
        ws.cell(row=row_num, column=3, value=log.operator.get_role_display() if log.operator else '-')
        ws.cell(row=row_num, column=4, value=log.action)
        ws.cell(row=row_num, column=5, value=log.module)
        ws.cell(row=row_num, column=6, value=log.description)
        ws.cell(row=row_num, column=7, value=str(log.ip_address))
        ws.cell(row=row_num, column=8, value=log.user_agent or '')

        # 设置数据行样式
        for col_num in range(1, 9):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin_border

    # 准备响应
    filename = f'操作日志_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response


@login_required
def payment_config_list(request):
    """支付管理 - 微信支付配置"""
    from .models import WeChatPayConfig

    configs = WeChatPayConfig.objects.all()

    context = {
        'configs': configs,
        'page_title': '支付管理',
    }
    context.update(get_common_context())
    return render(request, 'admin/payment_config.html', context)


@login_required
def account_management_list(request):
    """账户管理 - 账户和权限管理"""
    from .models import Permission, RolePermission
    from django.db.models import Count

    users = User.objects.all()
    permissions = Permission.objects.all()
    role_permissions = RolePermission.objects.select_related('permission').all()

    # 按角色分组统计
    role_stats = {}
    for role_choice in User.role.field.choices:
        role_code = role_choice[0]
        role_name = role_choice[1]
        count = users.filter(role=role_code).count()
        role_stats[role_code] = {
            'name': role_name,
            'count': count,
            'code': role_code,
        }

    context = {
        'users': users,
        'permissions': permissions,
        'role_permissions': role_permissions,
        'role_stats': role_stats,
        'page_title': '账户管理',
    }
    context.update(get_common_context())
    return render(request, 'admin/account_management.html', context)


# ============================================
# 微信支付相关视图
# ============================================

@login_required
def wechat_pay_config_list(request):
    """微信支付配置管理"""
    from apps.core.models import WeChatPayConfig

    configs = WeChatPayConfig.objects.all()

    context = {
        'configs': configs,
        'page_title': '微信支付配置',
    }
    context.update(get_common_context())
    return render(request, 'admin/wechat_pay_config.html', context)


@login_required
def wechat_pay_create_order(request):
    """创建微信支付订单"""
    from django.http import JsonResponse
    from django.views.decorators.csrf import csrf_exempt
    from apps.payment.models import PaymentBill
    from apps.core.wechat_pay import WeChatPayService

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '只支持POST请求'})

    bill_id = request.POST.get('bill_id')
    if not bill_id:
        return JsonResponse({'success': False, 'error': '缺少账单ID'})

    try:
        bill = PaymentBill.objects.get(id=bill_id)

        # 获取微信支付配置
        service = WeChatPayService()

        # 获取客户端IP
        client_ip = request.META.get('REMOTE_ADDR', '127.0.0.1')

        # 创建支付订单
        result = service.create_payment_order(bill, client_ip)

        return JsonResponse(result)

    except PaymentBill.DoesNotExist:
        return JsonResponse({'success': False, 'error': '账单不存在'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
def wechat_pay_notify(request):
    """微信支付回调通知"""
    from django.http import HttpResponse
    from django.views.decorators.http import require_POST
    from apps.payment.models import PaymentBill, PaymentRecord
    from apps.core.wechat_pay import WeChatPayService
    from django.utils import timezone

    if request.method != 'POST':
        return HttpResponse('只支持POST请求', status=405)

    # 获取微信支付配置
    service = WeChatPayService()

    # 验证回调
    verification = service.verify_notify(request.body)
    if not verification['success']:
        return HttpResponse('FAIL', status=400)

    data = verification['data']

    # 查找订单
    out_trade_no = data.get('out_trade_no')
    transaction_id = data.get('transaction_id')

    # 从商户订单号中提取bill_id（格式：PROP{bill_id_hex}_{timestamp}）
    try:
        bill_id_hex = out_trade_no.split('_')[0].replace('PROP', '')
        bill_id = bill_id_hex[:32]
        bill = PaymentBill.objects.get(id=bill_id)

        # 更新账单状态
        if data.get('result_code') == 'SUCCESS':
            # 支付成功
            total_fee = Decimal(data.get('total_fee', '0')) / 100  # 转换为元

            # 创建缴费记录
            PaymentRecord.objects.create(
                bill=bill,
                amount=total_fee,
                payment_method='wechat',
                operator='系统',
                payment_time=timezone.now(),
                transaction_id=transaction_id,
                remark=f'微信支付-{out_trade_no}'
            )

            return HttpResponse('SUCCESS')

    except Exception as e:
        print(f'处理支付回调异常: {e}')
        return HttpResponse('FAIL', status=500)


# ==================== 消息推送视图 ====================

@login_required
def notification_list(request):
    """
    获取用户通知列表

    GET参数:
        - unread_only: 是否只获取未读通知 (0/1)
        - limit: 返回数量限制 (默认50)
    """
    from .notification_service import NotificationService

    unread_only = request.GET.get('unread_only', '0') == '1'
    limit = int(request.GET.get('limit', 50))

    notifications = NotificationService.get_user_notifications(
        user_id=str(request.user.id),
        unread_only=unread_only,
        limit=limit
    )

    # 序列化通知数据
    notifications_data = []
    for notif in notifications:
        notifications_data.append({
            'id': str(notif.id),
            'type': notif.notification_type,
            'title': notif.title,
            'content': notif.content,
            'status': notif.status,
            'created_at': notif.created_at.isoformat(),
            'read_at': notif.read_at.isoformat() if notif.read_at else None,
            'related_bill_id': str(notif.related_bill_id) if notif.related_bill_id else None,
            'related_maintenance_id': str(notif.related_maintenance_id) if notif.related_maintenance_id else None,
        })

    return JsonResponse({
        'success': True,
        'notifications': notifications_data
    })


@login_required
def notification_mark_read(request, notification_id):
    """标记通知为已读"""
    try:
        notification = Notification.objects.get(id=notification_id, recipient=request.user)
        notification.mark_as_read()

        return JsonResponse({
            'success': True,
            'message': '已标记为已读'
        })
    except Notification.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': '通知不存在或无权访问'
        }, status=404)


@login_required
def notification_unread_count(request):
    """获取未读通知数量"""
    from .notification_service import NotificationService

    count = NotificationService.get_unread_count(str(request.user.id))

    return JsonResponse({
        'success': True,
        'unread_count': count
    })


@login_required
def notification_mark_all_read(request):
    """标记所有通知为已读"""
    from django.utils import timezone

    updated = Notification.objects.filter(
        recipient=request.user,
        status='sent'
    ).update(
        status='read',
        read_at=timezone.now()
    )

    return JsonResponse({
        'success': True,
        'message': f'已标记 {updated} 条通知为已读'
    })


@require_http_methods(["POST"])
@login_required
def send_test_notification(request):
    """发送测试通知（仅管理员）"""
    if request.user.role not in ['super_admin', 'admin']:
        return JsonResponse({
            'success': False,
            'error': '无权发送测试通知'
        }, status=403)

    from .notification_service import NotificationService

    import json
    data = json.loads(request.body)

    notification_type = data.get('type', 'system_announcement')
    title = data.get('title', '测试通知')
    content = data.get('content', '这是一条测试通知')

    try:
        if notification_type == 'system_announcement':
            notifications = NotificationService.send_system_announcement(title, content)
            return JsonResponse({
                'success': True,
                'message': f'已发送 {len(notifications)} 条系统公告'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': '不支持的测试通知类型'
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)
