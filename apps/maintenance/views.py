"""
Maintenance Views
"""
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.utils import timezone
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

from .models import MaintenanceRequest, MaintenanceLog
from .serializers import (MaintenanceRequestSerializer, MaintenanceRequestListSerializer,
                           MaintenanceCreateSerializer, MaintenanceAssignSerializer,
                           MaintenanceCompleteSerializer, MaintenanceLogSerializer)
from apps.core.permissions import IsReceptionistUser, IsAdminUser


class MaintenanceRequestViewSet(viewsets.ModelViewSet):
    """报事管理视图集"""
    queryset = MaintenanceRequest.objects.select_related('community', 'property').all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['community', 'property', 'category', 'status', 'priority']
    search_fields = ['request_number', 'reporter', 'description']
    ordering_fields = ['created_at', 'priority', 'status']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return MaintenanceRequestListSerializer
        elif self.action == 'create':
            return MaintenanceCreateSerializer
        return MaintenanceRequestSerializer

    def get_permissions(self):
        """创建和更新需要特定权限"""
        if self.action in ['create', 'assign', 'complete']:
            return [IsAuthenticated()]
        if self.action in ['update', 'partial_update', 'destroy']:
            return [IsReceptionistUser()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        """创建报事时自动设置社区"""
        property_obj = serializer.validated_data['property']
        serializer.save(community=property_obj.community)

    @action(detail=True, methods=['post'], permission_classes=[IsReceptionistUser])
    def assign(self, request, pk=None):
        """
        派单
        将报事指派给维修人员
        """
        maintenance = self.get_object()
        serializer = MaintenanceAssignSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        assigned_to = serializer.validated_data['assigned_to']

        # 更新状态
        maintenance.status = 'assigned'
        maintenance.assigned_to = assigned_to
        maintenance.assigned_at = timezone.now()
        maintenance.save()

        # 记录日志
        MaintenanceLog.objects.create(
            request=maintenance,
            operator=request.user.username,
            action='派单',
            description=f'已指派给 {assigned_to}'
        )

        # 发送通知
        from apps.core.notification_service import NotificationService
        NotificationService.send_maintenance_notification(maintenance, 'assigned')

        return Response({'message': f'已成功指派给 {assigned_to}'})

    @action(detail=True, methods=['post'], permission_classes=[IsReceptionistUser])
    def start(self, request, pk=None):
        """
        开始处理
        维修人员开始处理报事
        """
        maintenance = self.get_object()

        if maintenance.status not in ['assigned', 'pending']:
            return Response({'error': '当前状态不能开始处理'}, status=status.HTTP_400_BAD_REQUEST)

        maintenance.status = 'processing'
        maintenance.started_at = timezone.now()
        maintenance.save()

        # 记录日志
        MaintenanceLog.objects.create(
            request=maintenance,
            operator=request.user.username,
            action='开始处理',
            description=f'维修人员 {maintenance.assigned_to} 开始处理'
        )

        # 发送通知
        from apps.core.notification_service import NotificationService
        NotificationService.send_maintenance_notification(maintenance, 'processing')

        return Response({'message': '已开始处理'})

    @action(detail=True, methods=['post'], permission_classes=[IsReceptionistUser])
    def complete(self, request, pk=None):
        """
        完成报事
        维修完成，上传结果图片
        """
        maintenance = self.get_object()

        if maintenance.status != 'processing':
            return Response({'error': '当前状态不能完成'}, status=status.HTTP_400_BAD_REQUEST)

        serializer = MaintenanceCompleteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        maintenance.status = 'completed'
        maintenance.completed_at = timezone.now()
        maintenance.result_description = serializer.validated_data['result_description']
        maintenance.result_images = serializer.validated_data.get('result_images', [])
        maintenance.save()

        # 记录日志
        MaintenanceLog.objects.create(
            request=maintenance,
            operator=request.user.username,
            action='完成',
            description=serializer.validated_data['result_description'],
            images=serializer.validated_data.get('result_images', [])
        )

        # 发送通知
        from apps.core.notification_service import NotificationService
        NotificationService.send_maintenance_notification(maintenance, 'completed')

        return Response({'message': '报事已完成'})

    @action(detail=True, methods=['post'])
    def rate(self, request, pk=None):
        """
        评价
        业主对维修服务进行评价
        """
        maintenance = self.get_object()

        if maintenance.status != 'completed':
            return Response({'error': '只能对已完成的报事进行评价'}, status=status.HTTP_400_BAD_REQUEST)

        rating = request.data.get('rating')
        feedback = request.data.get('feedback', '')

        if not rating or not (1 <= int(rating) <= 5):
            return Response({'error': '评分必须在1-5之间'}, status=status.HTTP_400_BAD_REQUEST)

        maintenance.rating = rating
        maintenance.feedback = feedback
        maintenance.save()

        return Response({'message': '评价成功'})

    @action(detail=False, methods=['get'])
    def my_requests(self, request):
        """获取我的报事记录（业主/租户）"""
        user = request.user

        if user.role == 'owner':
            from apps.property.models import Owner
            try:
                owner = Owner.objects.get(user=user)
                # 获取业主所有房产的报事记录
                properties = [op.property for op in owner.owners.all()]
                requests = MaintenanceRequest.objects.filter(property__in=properties)
                serializer = MaintenanceRequestListSerializer(requests, many=True)
                return Response(serializer.data)
            except Owner.DoesNotExist:
                return Response({'error': '未找到业主信息'}, status=status.HTTP_404_NOT_FOUND)
        elif user.role == 'tenant':
            from apps.property.models import Tenant
            try:
                tenant = Tenant.objects.get(user=user, is_active=True)
                requests = MaintenanceRequest.objects.filter(property=tenant.property)
                serializer = MaintenanceRequestListSerializer(requests, many=True)
                return Response(serializer.data)
            except Tenant.DoesNotExist:
                return Response({'error': '未找到租户信息'}, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({'error': '无权访问'}, status=status.HTTP_403_FORBIDDEN)

    @action(detail=True, methods=['get'])
    def logs(self, request, pk=None):
        """获取报事的处理日志"""
        maintenance = self.get_object()
        logs = maintenance.logs.all()
        serializer = MaintenanceLogSerializer(logs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], permission_classes=[IsReceptionistUser])
    def close(self, request, pk=None):
        """
        关闭报事
        管理员或业主确认报事已完成并关闭
        """
        maintenance = self.get_object()

        if maintenance.status == 'closed':
            return Response({'error': '报事已关闭'}, status=status.HTTP_400_BAD_REQUEST)

        if maintenance.status not in ['completed', 'pending']:
            return Response({'error': '只能关闭已完成或待派单的报事'}, status=status.HTTP_400_BAD_REQUEST)

        old_status = maintenance.get_status_display()
        maintenance.status = 'closed'
        maintenance.save()

        # 记录日志
        MaintenanceLog.objects.create(
            request=maintenance,
            operator=request.user.username,
            action='关闭',
            description=f'报事从"{old_status}"状态关闭'
        )

        # 发送通知
        from apps.core.notification_service import NotificationService
        NotificationService.send_maintenance_notification(maintenance, 'closed')

        return Response({'message': '报事已关闭'})

    @action(detail=True, methods=['post'], permission_classes=[IsReceptionistUser])
    def reopen(self, request, pk=None):
        """
        重新打开报事
        将已完成的报事重新打开处理（返工）
        """
        maintenance = self.get_object()

        if maintenance.status != 'completed':
            return Response({'error': '只能重新打开已完成的报事'}, status=status.HTTP_400_BAD_REQUEST)

        maintenance.status = 'processing'
        maintenance.save()

        # 记录日志
        MaintenanceLog.objects.create(
            request=maintenance,
            operator=request.user.username,
            action='重新打开',
            description='报事重新打开处理（返工）'
        )

        # 发送通知
        from apps.core.notification_service import NotificationService
        NotificationService.send_maintenance_notification(maintenance, 'processing')

        return Response({'message': '报事已重新打开'})

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """导出报事数据到Excel"""
        # 获取筛选后的数据
        queryset = self.filter_queryset(self.get_queryset())

        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '报事记录'

        # 定义表头
        headers = [
            '报事编号', '小区', '房产地址', '报事类型', '优先级',
            '状态', '报事人', '联系电话', '详细描述',
            '指派给', '处理结果', '创建时间', '更新时间'
        ]

        # 设置表头样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 写入表头
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 设置列宽
        column_widths = {
            'A': 15, 'B': 15, 'C': 25, 'D': 12, 'E': 10,
            'F': 12, 'G': 12, 'H': 15, 'I': 40,
            'J': 12, 'K': 30, 'L': 18, 'M': 18
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        # 写入数据
        data_font = Font(size=10)
        data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for row_num, maintenance in enumerate(queryset, 2):
            # 获取最新的处理日志
            latest_log = maintenance.logs.order_by('-created_at').first()
            result_description = latest_log.description if latest_log and latest_log.action in ['完成', '关闭'] else ''

            # 状态映射
            status_map = {
                'pending': '待派单',
                'assigned': '已派单',
                'processing': '处理中',
                'completed': '已完成',
                'closed': '已关闭'
            }

            # 优先级映射
            priority_map = {
                'low': '低',
                'medium': '中',
                'high': '高',
                'urgent': '紧急'
            }

            # 报事类型映射
            category_map = {
                'plumbing': '水暖',
                'electrical': '电气',
                'structural': '结构',
                'appliance': '家电',
                'other': '其他'
            }

            data = [
                maintenance.request_number,
                maintenance.community.name if maintenance.community else '',
                str(maintenance.property) if maintenance.property else '',
                category_map.get(maintenance.category, maintenance.category),
                priority_map.get(maintenance.priority, maintenance.priority),
                status_map.get(maintenance.status, maintenance.status),
                maintenance.reporter or '',
                maintenance.reporter_phone or '',
                maintenance.description or '',
                maintenance.assigned_to or '',
                result_description,
                maintenance.created_at.strftime('%Y-%m-%d %H:%M:%S') if maintenance.created_at else '',
                maintenance.updated_at.strftime('%Y-%m-%d %H:%M:%S') if maintenance.updated_at else ''
            ]

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = data_font
                cell.alignment = data_alignment

        # 设置边框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 为所有数据单元格添加边框
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 生成文件名
        filename = f'报事记录_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        # 设置响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response


class MaintenanceLogViewSet(viewsets.ReadOnlyModelViewSet):
    """报事处理日志视图集"""
    queryset = MaintenanceLog.objects.select_related('request').all()
    serializer_class = MaintenanceLogSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['request', 'action']
    search_fields = ['operator', 'description']
    ordering = ['-created_at']

# ============================================
# Form Views for Admin Interface
# ============================================

@csrf_exempt
@login_required
def maintenance_request_form(request, pk=None):
    """渲染报事单表单"""
    from .forms import MaintenanceRequestForm

    if pk:
        maintenance_request = get_object_or_404(MaintenanceRequest, pk=pk)
        form = MaintenanceRequestForm(request.POST or None, instance=maintenance_request)
    else:
        form = MaintenanceRequestForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            # 保存时自动设置社区（从房产获取）
            maintenance_request = form.save(commit=False)
            # 使用 property_id 直接获取社区，避免 RelatedObjectDoesNotExist 异常
            if maintenance_request.property_id:
                from apps.property.models import Property
                property_obj = Property.objects.get(id=maintenance_request.property_id)
                maintenance_request.community = property_obj.community
            maintenance_request.save()
            return JsonResponse({
                'success': True,
                'message': '保存成功',
                'data': {
                    'id': maintenance_request.id,
                    'request_number': maintenance_request.request_number,
                    'status': maintenance_request.status
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': dict([(field, [str(e) for e in errors]) for field, errors in form.errors.items()])
            })

    from django.template.loader import render_to_string
    from apps.community.models import Community

    communities = Community.objects.all().order_by('name')
    form_url = request.path if pk is None else request.path

    form_html = render_to_string('admin/forms/maintenance_request_form.html', {
        'form': form,
        'csrf_token': request.META.get('CSRF_COOKIE', ''),
        'communities': communities,
        'form_url': form_url
    })

    return JsonResponse({'html': form_html})
