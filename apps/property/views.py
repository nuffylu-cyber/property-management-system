"""
Property Views
"""
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.shortcuts import get_object_or_404
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
import pandas as pd
from django.db import transaction

from .models import Property, Owner, Tenant, OwnerProperty
from .serializers import (PropertySerializer, PropertyListSerializer, OwnerSerializer,
                           OwnerListSerializer, TenantSerializer, OwnerPropertyRelationSerializer)
from apps.core.permissions import IsAdminUser, IsReceptionistUser
from apps.community.models import Community, Building


class PropertyViewSet(viewsets.ModelViewSet):
    """房产管理视图集"""
    queryset = Property.objects.select_related('community', 'building').all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['community', 'building', 'property_type', 'status']
    search_fields = ['room_number']
    ordering_fields = ['community', 'building', 'floor', 'room_number']
    ordering = ['community', 'building', 'floor', 'room_number']

    def get_serializer_class(self):
        if self.action == 'list':
            return PropertyListSerializer
        return PropertySerializer

    def get_permissions(self):
        """只有管理员和前台可以创建/修改/删除"""
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return [IsAuthenticated()]

    @action(detail=False, methods=['get'])
    def my_properties(self, request):
        """获取我的房产列表（业主/租户）"""
        user = request.user
        if user.role == 'owner':
            try:
                owner = Owner.objects.get(user=user)
                properties = [prop.property for prop in owner.owners.all()]
                serializer = PropertyListSerializer(properties, many=True)
                return Response(serializer.data)
            except Owner.DoesNotExist:
                return Response({'error': '未找到业主信息'}, status=status.HTTP_404_NOT_FOUND)
        elif user.role == 'tenant':
            try:
                tenant = Tenant.objects.get(user=user, is_active=True)
                serializer = PropertyListSerializer([tenant.property], many=True)
                return Response(serializer.data)
            except Tenant.DoesNotExist:
                return Response({'error': '未找到租户信息'}, status=status.HTTP_404_NOT_FOUND)
        else:
            return Response({'error': '无权访问'}, status=status.HTTP_403_FORBIDDEN)

    @action(detail=False, methods=['post'], permission_classes=[IsAdminUser])
    def import_excel(self, request):
        """从Excel导入房产和业主信息（支持多个工作表）"""
        if 'file' not in request.FILES:
            return Response({'error': '请上传Excel文件'}, status=status.HTTP_400_BAD_REQUEST)

        excel_file = request.FILES['file']

        # 检查文件扩展名
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': '只支持.xlsx或.xls格式的文件'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # 优先从前端获取community_id参数
            community_id = request.data.get('community_id')

            if community_id:
                # 使用前端传递的小区ID
                try:
                    community = Community.objects.get(id=community_id)
                except Community.DoesNotExist:
                    return Response({
                        'error': f'所选小区不存在（ID: {community_id}）'
                    }, status=status.HTTP_400_BAD_REQUEST)
            else:
                # 备选方案：根据Excel文件名自动识别小区
                filename = excel_file.name
                all_communities = Community.objects.all()

                # 在文件名中查找匹配的小区名称
                matched_community = None
                for community in all_communities:
                    if community.name in filename:
                        matched_community = community
                        break

                # 如果没有匹配到小区，返回错误提示
                if not matched_community:
                    return Response({
                        'error': f'请先选择小区。文件名"{filename}"中没有找到匹配的小区名称。'
                    }, status=status.HTTP_400_BAD_REQUEST)

                community = matched_community

            # 读取Excel文件的所有工作表
            excel_file_obj = pd.ExcelFile(excel_file)
            sheet_names = excel_file_obj.sheet_names

            # 总体统计
            total_stats = {
                'sheets_processed': 0,
                'total_sheets': len(sheet_names),
                'created_properties': 0,
                'updated_properties': 0,
                'created_owners': 0,
                'linked_owners': 0,
                'errors': [],
                'sheet_details': []
            }

            # 遍历所有工作表
            for sheet_name in sheet_names:
                # 读取当前工作表
                # 根据文件扩展名选择引擎
                if excel_file.name.endswith('.xls'):
                    df = pd.read_excel(excel_file_obj, sheet_name=sheet_name, engine='xlrd')
                else:
                    df = pd.read_excel(excel_file_obj, sheet_name=sheet_name)

                # 检测Excel格式
                # 格式1: 旧格式 - 列名直接包含'房号', '姓名', '面积', '电话号码'
                # 格式2: 新格式 - 列名在第二行(header=1)，包含'小区', '楼栋', '单元', '楼层', '房号', '业主姓名', '联系电话'

                is_old_format = '房号' in df.columns and '姓名' in df.columns

                if not is_old_format:
                    # 尝试新格式：读取第二行作为表头
                    try:
                        # 根据文件扩展名选择引擎
                        if excel_file.name.endswith('.xls'):
                            df_new = pd.read_excel(excel_file_obj, sheet_name=sheet_name, header=1, engine='xlrd')
                        else:
                            df_new = pd.read_excel(excel_file_obj, sheet_name=sheet_name, header=1)
                        # 检查是否包含新格式的列
                        # 新格式的列可能是Unnamed，但数据中包含小区、楼栋等信息
                        if len(df_new.columns) >= 8:
                            # 使用新格式
                            df = df_new
                            is_old_format = False
                        else:
                            continue
                    except:
                        continue
                else:
                    # 旧格式，支持两种列名：电话号码/电话
                    if '电话' in df.columns and '电话号码' not in df.columns:
                        df.rename(columns={'电话': '电话号码'}, inplace=True)

                # 处理这个工作表
                sheet_stats = {
                    'sheet_name': sheet_name,
                    'rows_total': len(df),
                    'created_properties': 0,
                    'updated_properties': 0,
                    'created_owners': 0,
                    'linked_owners': 0,
                    'errors': []
                }

                with transaction.atomic():
                    if is_old_format:
                        # 旧格式处理逻辑
                        current_room_number = None
                        current_area = None
                        current_property = None

                        for idx, row in df.iterrows():
                            try:
                                # 处理房号（空房号表示多业主情况）
                                room_str = str(row['房号']).strip() if pd.notna(row['房号']) else None

                                if room_str and room_str != 'nan':
                                    current_room_number = room_str
                                    current_area = float(row['面积']) if pd.notna(row['面积']) else None

                                # 如果房号为空且没有当前房产，跳过
                                if not current_room_number:
                                    sheet_stats['errors'].append(f'第{idx+2}行: 缺少房号信息')
                                    continue

                                # 解析房号 (1-501 -> 1栋, 5层, 01号房)
                                parts = current_room_number.split('-')
                                if len(parts) != 2:
                                    sheet_stats['errors'].append(f'第{idx+2}行: 房号格式错误 {current_room_number}')
                                    continue

                                building_num = parts[0]
                                floor_room = parts[1]

                                if not floor_room:
                                    sheet_stats['errors'].append(f'第{idx+2}行: 房号格式错误 {current_room_number}')
                                    continue

                                # 解析楼层和房号
                                if len(floor_room) == 3:
                                    floor = int(floor_room[0])
                                    room_number = floor_room[1:3]
                                elif len(floor_room) >= 4:
                                    floor = int(floor_room[:-2])
                                    room_number = floor_room[-2:]
                                else:
                                    sheet_stats['errors'].append(f'第{idx+2}行: 房号格式错误 {current_room_number}')
                                    continue

                                # 获取或创建楼栋
                                building, _ = Building.objects.get_or_create(
                                    community=community,
                                    name=f'{building_num}号楼',
                                    defaults={'description': f'{building_num}号楼'}
                                )

                                # 获取或创建房产
                                area = float(row['面积']) if pd.notna(row['面积']) else current_area
                                if not area:
                                    sheet_stats['errors'].append(f'第{idx+2}行: 缺少面积信息')
                                    continue

                                property_obj, created = Property.objects.get_or_create(
                                    building=building,
                                    floor=floor,
                                    room_number=room_number,
                                    defaults={
                                        'community': community,
                                        'area': area,
                                        'property_type': 'residential',
                                        'status': 'occupied'
                                    }
                                )

                                if created:
                                    sheet_stats['created_properties'] += 1
                                else:
                                    if property_obj.area != area:
                                        property_obj.area = area
                                        property_obj.save()
                                        sheet_stats['updated_properties'] += 1

                                current_property = property_obj

                                # 处理业主信息
                                owner_name = str(row['姓名']).strip() if pd.notna(row['姓名']) else None
                                if not owner_name or owner_name == 'nan':
                                    continue

                                owner_phone = str(row['电话号码']).strip() if pd.notna(row['电话号码']) else None
                                if not owner_phone or owner_phone == 'nan':
                                    sheet_stats['errors'].append(f'第{idx+2}行: 缺少电话号码')
                                    continue

                                # 清理电话号码
                                try:
                                    if isinstance(owner_phone, float) or '.' in owner_phone:
                                        owner_phone = str(int(float(owner_phone)))
                                except:
                                    pass

                                # 获取或创建业主
                                owner, owner_created = Owner.objects.get_or_create(
                                    phone=owner_phone,
                                    defaults={
                                        'name': owner_name,
                                        'is_verified': True
                                    }
                                )

                                if owner_created:
                                    sheet_stats['created_owners'] += 1
                                else:
                                    # 更新姓名
                                    if owner.name != owner_name:
                                        owner.name = owner_name
                                        owner.save()

                                # 关联业主和房产
                                owner_property, rel_created = OwnerProperty.objects.get_or_create(
                                    property=property_obj,
                                    owner=owner,
                                    defaults={'is_primary': True}
                                )

                                if rel_created:
                                    sheet_stats['linked_owners'] += 1

                            except Exception as e:
                                sheet_stats['errors'].append(f'第{idx+2}行: {str(e)}')

                    else:
                        # 新格式处理逻辑
                        # 列映射: 根据实际Excel结构调整
                        for idx, row in df.iterrows():
                            try:
                                # 跳过空行或标题行
                                if pd.isna(row.iloc[1]):  # 小区列为空
                                    continue

                                # 提取数据
                                community_name = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else None
                                building_full_name = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else None
                                unit = str(row.iloc[3]).strip() if pd.notna(row.iloc[3]) else None
                                floor = row.iloc[4]
                                room_number_raw = row.iloc[5]
                                owner_name = str(row.iloc[7]).strip() if pd.notna(row.iloc[7]) else None
                                owner_phone = str(row.iloc[8]).strip() if pd.notna(row.iloc[8]) else None
                                area = row.iloc[9]

                                # 验证必需字段
                                if not community_name or community_name == 'nan':
                                    continue
                                if not building_full_name or building_full_name == 'nan':
                                    continue
                                if not owner_name or owner_name == 'nan':
                                    continue
                                if not owner_phone or owner_phone == 'nan':
                                    continue
                                if pd.isna(floor) or floor == 'nan':
                                    continue
                                if pd.isna(room_number_raw):
                                    continue
                                if pd.isna(area):
                                    continue

                                # 清理数据
                                try:
                                    floor = int(floor)
                                    area = float(area)
                                except:
                                    sheet_stats['errors'].append(f'第{idx+2}行: 数据格式错误')
                                    continue

                                # 清理房号：处理数字类型（如3007、3007.0）
                                if isinstance(room_number_raw, (int, float)):
                                    # 如果是整数或浮点数，转换为整数再转字符串
                                    try:
                                        room_number = str(int(room_number_raw))
                                        # 确保至少有2位（例如：7 -> "07"）
                                        if len(room_number) == 1:
                                            room_number = '0' + room_number
                                    except:
                                        room_number = str(room_number_raw).strip()
                                else:
                                    room_number = str(room_number_raw).strip()

                                # 移除房号中的小数点（如果有的话）
                                if '.' in room_number:
                                    room_number = room_number.split('.')[0]

                                # 清理电话号码
                                try:
                                    if isinstance(owner_phone, float) or '.' in str(owner_phone):
                                        owner_phone = str(int(float(owner_phone)))
                                except:
                                    pass

                                # 从完整楼栋名称中提取楼栋号
                                # 例如: "锦尚名都-1号楼" -> "1号楼"
                                if '-' in building_full_name:
                                    building_name = building_full_name.split('-')[-1]
                                    if not building_name.endswith('号楼'):
                                        building_name = f'{building_name}号楼'
                                else:
                                    building_name = building_full_name

                                # 获取或创建楼栋
                                building, _ = Building.objects.get_or_create(
                                    community=community,
                                    name=building_name,
                                    defaults={'description': building_name}
                                )

                                # 获取或创建房产（包含单元作为唯一键）
                                property_obj, created = Property.objects.get_or_create(
                                    building=building,
                                    unit=unit,
                                    floor=floor,
                                    room_number=room_number,
                                    defaults={
                                        'community': community,
                                        'area': area,
                                        'property_type': 'residential',
                                        'status': 'occupied'
                                    }
                                )

                                if created:
                                    sheet_stats['created_properties'] += 1
                                else:
                                    # 更新字段（unit已在查找条件中，无需更新）
                                    updated = False
                                    if property_obj.area != area:
                                        property_obj.area = area
                                        updated = True
                                    if updated:
                                        property_obj.save()
                                        sheet_stats['updated_properties'] += 1

                                # 获取或创建业主
                                owner, owner_created = Owner.objects.get_or_create(
                                    phone=owner_phone,
                                    defaults={
                                        'name': owner_name,
                                        'is_verified': True
                                    }
                                )

                                if owner_created:
                                    sheet_stats['created_owners'] += 1
                                else:
                                    # 更新姓名
                                    if owner.name != owner_name:
                                        owner.name = owner_name
                                        owner.save()

                                # 关联业主和房产
                                owner_property, rel_created = OwnerProperty.objects.get_or_create(
                                    property=property_obj,
                                    owner=owner,
                                    defaults={'is_primary': True}
                                )

                                if rel_created:
                                    sheet_stats['linked_owners'] += 1

                            except Exception as e:
                                sheet_stats['errors'].append(f'第{idx+2}行: {str(e)}')
                                continue


                # 累加到总统计
                total_stats['created_properties'] += sheet_stats['created_properties']
                total_stats['updated_properties'] += sheet_stats['updated_properties']
                total_stats['created_owners'] += sheet_stats['created_owners']
                total_stats['linked_owners'] += sheet_stats['linked_owners']
                total_stats['errors'].extend([f"[{sheet_name}] {err}" for err in sheet_stats['errors']])
                total_stats['sheets_processed'] += 1
                total_stats['sheet_details'].append(sheet_stats)

            return Response({
                'message': f'Excel导入完成，共处理 {total_stats["sheets_processed"]} 个工作表',
                'stats': total_stats
            })

        except Exception as e:
            return Response({
                'error': f'导入失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], permission_classes=[IsAdminUser])
    def export_excel(self, request):
        """导出房产列表到Excel（支持筛选）"""
        from django.http import HttpResponse
        from django.db.models import Q, Prefetch
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        try:
            # 获取筛选参数
            community_id = request.query_params.get('community', '')
            property_type = request.query_params.get('property_type', '')
            status_filter = request.query_params.get('status', '')
            search_query = request.query_params.get('search', '')

            # 构建查询集
            queryset = Property.objects.select_related(
                'community', 'building'
            ).prefetch_related(
                Prefetch(
                    'owners',
                    queryset=OwnerProperty.objects.select_related('owner').order_by('-is_primary', '-created_at')
                )
            )

            # 应用筛选条件
            if community_id:
                queryset = queryset.filter(community_id=community_id)
            if property_type:
                queryset = queryset.filter(property_type=property_type)
            if status_filter:
                queryset = queryset.filter(status=status_filter)

            # 应用搜索逻辑
            if search_query:
                # 尝试解析数字房号
                floor_room_match = None
                if search_query.isdigit() and len(search_query) >= 3:
                    query_len = len(search_query)
                    if query_len == 3:
                        floor_num = int(search_query[0])
                        room_suffix = search_query[1:]
                        floor_room_match = (floor_num, room_suffix)
                    elif query_len == 4:
                        floor_num = int(search_query[:2])
                        room_suffix = search_query[2:]
                        floor_room_match = (floor_num, room_suffix)
                    elif query_len == 5:
                        floor_num = int(search_query[:3])
                        room_suffix = search_query[3:]
                        floor_room_match = (floor_num, room_suffix)

                queries = Q(building__name__icontains=search_query) | Q(room_number__icontains=search_query)
                if floor_room_match:
                    floor_num, room_suffix = floor_room_match
                    queries |= Q(floor=floor_num, room_number__startswith=room_suffix)

                queryset = queryset.filter(queries)

                # 搜索业主
                owner_ids = Owner.objects.filter(
                    Q(name__icontains=search_query) | Q(phone__icontains=search_query)
                ).values_list('id', flat=True)

                if owner_ids:
                    property_ids_from_owners = OwnerProperty.objects.filter(
                        owner_id__in=owner_ids
                    ).values_list('property_id', flat=True)

                    queryset = queryset | Property.objects.filter(
                        id__in=property_ids_from_owners
                    ).select_related(
                        'community', 'building'
                    ).prefetch_related(
                        Prefetch(
                            'owners',
                            queryset=OwnerProperty.objects.select_related('owner').order_by('-is_primary', '-created_at')
                        )
                    )

            queryset = queryset.order_by('building', 'floor', 'room_number').distinct()

            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "房产列表"

            # 定义样式
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            border_style = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            # 设置列宽
            ws.column_dimensions['A'].width = 15  # 房号
            ws.column_dimensions['B'].width = 20  # 小区
            ws.column_dimensions['C'].width = 10  # 楼层
            ws.column_dimensions['D'].width = 10  # 面积
            ws.column_dimensions['E'].width = 10  # 类型
            ws.column_dimensions['F'].width = 15  # 业主
            ws.column_dimensions['G'].width = 15  # 联系电话
            ws.column_dimensions['H'].width = 10  # 状态

            # 写入表头
            headers = ['房号', '所属小区', '楼层', '面积(㎡)', '类型', '业主', '联系电话', '状态']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border_style

            # 写入数据
            row_num = 2
            for property_obj in queryset:
                # 获取业主信息
                owners_list = property_obj.owners.all()[:2]  # 最多显示2个业主
                owner_names = '、'.join([op.owner.name for op in owners_list])
                owner_phones = '、'.join([op.owner.phone for op in owners_list])

                # 房号
                ws.cell(row=row_num, column=1, value=str(property_obj))
                # 小区
                ws.cell(row=row_num, column=2, value=property_obj.community.name)
                # 楼层
                ws.cell(row=row_num, column=3, value=property_obj.floor)
                # 面积
                ws.cell(row=row_num, column=4, value=float(property_obj.area))
                # 类型
                property_type_map = {'residential': '住宅', 'commercial': '商业', 'garage': '车库', 'storage': '储藏室'}
                ws.cell(row=row_num, column=5, value=property_type_map.get(property_obj.property_type, property_obj.property_type))
                # 业主
                ws.cell(row=row_num, column=6, value=owner_names)
                # 联系电话
                ws.cell(row=row_num, column=7, value=owner_phones)
                # 状态
                status_map = {'occupied': '自住', 'rented': '出租', 'vacant': '空置', 'renovation': '装修中'}
                ws.cell(row=row_num, column=8, value=status_map.get(property_obj.status, property_obj.status))

                # 设置单元格样式和边框
                for col in range(1, 9):
                    cell = ws.cell(row=row_num, column=col)
                    cell.alignment = cell_alignment
                    cell.border = border_style

                row_num += 1

            # 准备响应
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="房产列表_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx"'

            # 保存到BytesIO
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            response.write(output.read())

            return response

        except Exception as e:
            return Response({
                'error': f'导出失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['delete'], permission_classes=[IsAdminUser])
    def batch_delete(self, request):
        """批量删除房产"""
        property_ids = request.data.get('property_ids', [])

        if not property_ids:
            return Response({
                'error': '请选择要删除的房产'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            # 获取要删除的房产
            properties = Property.objects.filter(id__in=property_ids)
            count = properties.count()

            if count == 0:
                return Response({
                    'error': '未找到要删除的房产'
                }, status=status.HTTP_404_NOT_FOUND)

            # 删除房产（级联删除业主关联关系）
            properties.delete()

            return Response({
                'success': True,
                'message': f'成功删除 {count} 个房产'
            })

        except Exception as e:
            return Response({
                'error': f'批量删除失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class OwnerViewSet(viewsets.ModelViewSet):
    """业主管理视图集"""
    queryset = Owner.objects.all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_verified']
    search_fields = ['name', 'phone', 'id_card']
    ordering_fields = ['created_at']
    ordering = ['-created_at']

    def get_serializer_class(self):
        if self.action == 'list':
            return OwnerListSerializer
        return OwnerSerializer

    def get_permissions(self):
        """只有管理员和前台可以创建/修改/删除"""
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return [IsAuthenticated()]

    @action(detail=True, methods=['get'])
    def properties(self, request, pk=None):
        """获取业主的所有房产"""
        owner = self.get_object()
        properties = [prop.property for prop in owner.owners.all()]
        serializer = PropertyListSerializer(properties, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'], permission_classes=[IsAdminUser])
    def verify(self, request, pk=None):
        """验证业主身份"""
        owner = self.get_object()
        owner.is_verified = True
        owner.save()
        return Response({'message': '业主身份验证成功'})

    @action(detail=True, methods=['post'], permission_classes=[IsAdminUser])
    def bind_wechat(self, request, pk=None):
        """绑定微信"""
        owner = self.get_object()
        wechat_openid = request.data.get('wechat_openid')
        if not wechat_openid:
            return Response({'error': '缺少 wechat_openid'}, status=status.HTTP_400_BAD_REQUEST)

        owner.wechat_openid = wechat_openid
        owner.wechat_nickname = request.data.get('wechat_nickname', '')
        owner.avatar_url = request.data.get('avatar_url', '')
        owner.save()
        return Response({'message': '微信绑定成功'})

    @action(detail=False, methods=['delete'], permission_classes=[IsAdminUser])
    def batch_delete(self, request):
        """批量删除业主"""
        owner_ids = request.data.get('owner_ids', [])

        if not owner_ids:
            return Response({
                'error': '请选择要删除的业主'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            # 获取要删除的业主
            owners = Owner.objects.filter(id__in=owner_ids)
            count = owners.count()

            if count == 0:
                return Response({
                    'error': '未找到要删除的业主'
                }, status=status.HTTP_404_NOT_FOUND)

            # 删除业主（级联删除房产关联关系）
            owners.delete()

            return Response({
                'success': True,
                'message': f'成功删除 {count} 位业主'
            })

        except Exception as e:
            return Response({
                'error': f'批量删除失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class TenantViewSet(viewsets.ModelViewSet):
    """租户管理视图集"""
    queryset = Tenant.objects.select_related('property').all()
    serializer_class = TenantSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['property', 'is_active']
    search_fields = ['name', 'phone']
    ordering_fields = ['created_at', 'lease_start', 'lease_end']
    ordering = ['-created_at']

    def get_permissions(self):
        """只有管理员和前台可以创建/修改/删除"""
        if self.action in ['create', 'update', 'partial_update', 'destroy']:
            return [IsAdminUser()]
        return [IsAuthenticated()]

# ============================================
# Form Views for Admin Interface
# ============================================

@csrf_exempt
@login_required
def property_form(request, pk=None):
    """渲染房产表单（支持业主信息同步）"""
    from .forms import PropertyForm
    from django.db import transaction

    if pk:
        property_obj = get_object_or_404(Property, pk=pk)
        form = PropertyForm(request.POST or None, instance=property_obj)
    else:
        form = PropertyForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            try:
                with transaction.atomic():
                    # 保存房产信息
                    property_obj = form.save()

                    # 处理多个业主
                    owners_count = int(request.POST.get('owners_count', 1))

                    # 收集所有业主数据
                    owners_data = []
                    for i in range(owners_count):
                        owner_name = request.POST.get(f'owner_name_{i}', '').strip()
                        owner_phone = request.POST.get(f'owner_phone_{i}', '').strip()
                        owner_ratio = request.POST.get(f'owner_ratio_{i}', '100')
                        owner_id = request.POST.get(f'owner_id_{i}', '')  # 用于标识已有业主

                        # 只有当填写了业主姓名时才处理
                        if owner_name:
                            owners_data.append({
                                'name': owner_name,
                                'phone': owner_phone,
                                'ratio': float(owner_ratio) if owner_ratio else 100.0,
                                'owner_id': owner_id
                            })

                    if not owners_data:
                        return JsonResponse({
                            'success': False,
                            'errors': {'__all__': ['至少需要一个业主']}
                        })

                    # 删除该房产的所有旧业主关联
                    OwnerProperty.objects.filter(property=property_obj).delete()

                    # 创建/更新业主并关联房产
                    for idx, owner_data in enumerate(owners_data):
                        # 查找或创建业主
                        owner = None
                        created = False

                        # 优先通过data-owner-id查找已存在的业主
                        owner_id = request.POST.get(f'owner_id_{idx}', '').strip()
                        if owner_id:
                            owner = Owner.objects.filter(id=owner_id).first()

                        # 如果没找到，通过电话号码查找
                        if not owner and owner_data['phone']:
                            owner = Owner.objects.filter(phone=owner_data['phone']).first()

                        # 如果还是没找到，通过姓名查找
                        if not owner:
                            owner = Owner.objects.filter(name=owner_data['name']).first()

                        # 如果都没找到，创建新业主
                        if not owner:
                            owner = Owner.objects.create(
                                name=owner_data['name'],
                                phone=owner_data['phone'] if owner_data['phone'] else '',
                                verification_status='approved',
                                source='admin_added',
                                is_verified=True
                            )
                            created = True
                        else:
                            # 找到了业主，更新信息
                            needs_save = False
                            if owner.name != owner_data['name']:
                                owner.name = owner_data['name']
                                needs_save = True
                            if owner.phone != owner_data['phone']:
                                owner.phone = owner_data['phone']
                                needs_save = True
                            if needs_save:
                                owner.save()

                        # 关联房产和业主
                        OwnerProperty.objects.create(
                            owner=owner,
                            property=property_obj,
                            ownership_type='full',
                            ownership_ratio=owner_data['ratio'],
                            is_primary=(idx == 0)  # 第一个业主为主要业主
                        )

            except Exception as e:
                import traceback
                return JsonResponse({
                    'success': False,
                    'errors': {'__all__': [str(e)]}
                })

            return JsonResponse({
                'success': True,
                'message': '保存成功',
                'data': {
                    'id': str(property_obj.id),
                    'room_number': property_obj.room_number,
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': dict([(field, [str(e) for e in errors]) for field, errors in form.errors.items()])
            })

    from django.template.loader import render_to_string
    form_html = render_to_string('admin/forms/property_form.html', {
        'form': form,
        'csrf_token': request.META.get('CSRF_COOKIE', '')
    })

    return JsonResponse({'html': form_html})


@csrf_exempt
@login_required
def get_properties_by_community(request):
    """获取指定小区的房产列表（用于级联选择）"""
    community_id = request.GET.get('community_id')

    if not community_id:
        return JsonResponse({'properties': []})

    try:
        # 获取该小区的所有房产
        properties = Property.objects.filter(
            community_id=community_id
        ).select_related(
            'community', 'building'
        ).order_by('building__name', 'unit', 'floor', 'room_number')

        # 构建房产列表数据
        properties_data = []
        for prop in properties:
            properties_data.append({
                'id': str(prop.id),
                'name': str(prop),
                'full_address': prop.full_address
            })

        return JsonResponse({'properties': properties_data})

    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'properties': []
        }, status=400)



@csrf_exempt
@login_required
def owner_form(request, pk=None):
    """渲染业主表单"""
    from .forms import OwnerForm

    if pk:
        owner = get_object_or_404(Owner, pk=pk)
        form = OwnerForm(request.POST or None, instance=owner)
    else:
        form = OwnerForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            try:
                with transaction.atomic():
                    # 1. 保存业主基本信息
                    owner = form.save()

                    # 2. 处理房产绑定
                    properties_count = int(request.POST.get('properties_count', 1))

                    # 收集所有房产数据
                    properties_data = []
                    property_ids = set()  # 用于检测重复

                    for i in range(properties_count):
                        property_id = request.POST.get(f'property_id_{i}', '').strip()
                        property_ratio = request.POST.get(f'property_ratio_{i}', '100')
                        op_id = request.POST.get(f'property_op_id_{i}', '')  # OwnerProperty ID

                        # 只有当选择了房产时才处理
                        if property_id:
                            # 防止重复绑定同一套房产
                            if property_id in property_ids:
                                return JsonResponse({
                                    'success': False,
                                    'errors': {'__all__': [f'不能重复绑定同一套房产 (ID: {property_id})']}
                                })
                            property_ids.add(property_id)

                            properties_data.append({
                                'property_id': property_id,
                                'ratio': float(property_ratio) if property_ratio else 100.0,
                                'op_id': op_id
                            })

                    # 3. 清理旧关联(只删除当前提交中不包含的)
                    # 获取当前业主的所有房产ID
                    existing_property_ids = set(
                        OwnerProperty.objects.filter(owner=owner)
                        .values_list('property_id', flat=True)
                    )

                    # 提交的新房产ID集合
                    new_property_ids = set(pd['property_id'] for pd in properties_data)

                    # 删除不再需要的关联
                    to_delete = existing_property_ids - new_property_ids
                    if to_delete:
                        OwnerProperty.objects.filter(
                            owner=owner,
                            property_id__in=to_delete
                        ).delete()

                    # 4. 创建或更新房产关联
                    for prop_data in properties_data:
                        property_obj = Property.objects.get(id=prop_data['property_id'])

                        # 检查是否已存在关联
                        owner_property = OwnerProperty.objects.filter(
                            owner=owner,
                            property=property_obj
                        ).first()

                        if owner_property:
                            # 更新已有关联
                            owner_property.ownership_ratio = prop_data['ratio']
                            owner_property.save()
                        else:
                            # 创建新关联
                            OwnerProperty.objects.create(
                                owner=owner,
                                property=property_obj,
                                ownership_type='full',
                                ownership_ratio=prop_data['ratio'],
                                is_primary=False  # 业主的所有房产平等,不设主要房产
                            )

                return JsonResponse({
                    'success': True,
                    'message': '保存成功',
                    'data': {
                        'id': str(owner.id),
                        'name': owner.name,
                        'phone': owner.phone
                    }
                })

            except Exception as e:
                import traceback
                traceback.print_exc()
                return JsonResponse({
                    'success': False,
                    'errors': {'__all__': [str(e)]}
                })
        else:
            return JsonResponse({
                'success': False,
                'errors': dict([(field, [str(e) for e in errors]) for field, errors in form.errors.items()])
            })

    # GET请求:渲染表单
    from django.template.loader import render_to_string

    # 获取所有小区列表(用于房产选择)
    communities = Community.objects.all().order_by('name')

    form_html = render_to_string('admin/forms/owner_form.html', {
        'form': form,
        'csrf_token': request.META.get('CSRF_COOKIE', ''),
        'communities': communities  # 传递给模板
    })

    return JsonResponse({'html': form_html})


@csrf_exempt
@login_required
def get_properties_by_community(request):
    """获取指定小区的房产列表（用于级联选择）"""
    community_id = request.GET.get('community_id')

    if not community_id:
        return JsonResponse({'properties': []})

    try:
        # 获取该小区的所有房产
        properties = Property.objects.filter(
            community_id=community_id
        ).select_related(
            'community', 'building'
        ).order_by('building__name', 'unit', 'floor', 'room_number')

        # 构建房产列表数据
        properties_data = []
        for prop in properties:
            properties_data.append({
                'id': str(prop.id),
                'name': str(prop),
                'full_address': prop.full_address
            })

        return JsonResponse({'properties': properties_data})

    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'properties': []
        }, status=400)



@csrf_exempt
@login_required
def tenant_form(request, pk=None):
    """渲染租户表单"""
    from .forms import TenantForm
    from apps.community.models import Community

    if pk:
        tenant = get_object_or_404(Tenant, pk=pk)
        form = TenantForm(request.POST or None, instance=tenant)
    else:
        form = TenantForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            tenant = form.save()
            return JsonResponse({
                'success': True,
                'message': '保存成功',
                'data': {
                    'id': tenant.id,
                    'name': tenant.name,
                    'phone': tenant.phone
                }
            })
        else:
            return JsonResponse({
                'success': False,
                'errors': dict([(field, [str(e) for e in errors]) for field, errors in form.errors.items()])
            })

    from django.template.loader import render_to_string
    # 获取所有小区列表
    communities = Community.objects.all().order_by('name')

    # 获取当前租户的房产信息（用于编辑模式）
    tenant_property_id = None
    tenant_property_community_id = None
    if pk and hasattr(form.instance, 'property') and form.instance.property:
        tenant_property_id = str(form.instance.property.id)
        tenant_property_community_id = str(form.instance.property.community.id)

    form_html = render_to_string('admin/forms/tenant_form.html', {
        'form': form,
        'communities': communities,
        'tenant_property_id': tenant_property_id,
        'tenant_property_community_id': tenant_property_community_id,
        'csrf_token': request.META.get('CSRF_COOKIE', '')
    })

    return JsonResponse({'html': form_html})


@csrf_exempt
@login_required
def get_properties_by_community(request):
    """获取指定小区的房产列表（用于级联选择）"""
    community_id = request.GET.get('community_id')

    if not community_id:
        return JsonResponse({'properties': []})

    try:
        # 获取该小区的所有房产
        properties = Property.objects.filter(
            community_id=community_id
        ).select_related(
            'community', 'building'
        ).order_by('building__name', 'unit', 'floor', 'room_number')

        # 构建房产列表数据
        properties_data = []
        for prop in properties:
            properties_data.append({
                'id': str(prop.id),
                'name': str(prop),
                'full_address': prop.full_address
            })

        return JsonResponse({'properties': properties_data})

    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'properties': []
        }, status=400)

